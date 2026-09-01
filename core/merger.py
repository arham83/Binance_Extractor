#!/usr/bin/env python3
"""Merge preprocessed daily Binance kline CSVs into one file per year.

For Excel output, install:
    python -m pip install openpyxl

For Parquet output, also install:
    python -m pip install pandas pyarrow

Example:
    python merge_yearly_klines.py --format parquet
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from datetime import date, time
from pathlib import Path

FILENAME_PATTERN = re.compile(
    r"^(?P<coin>.+)-(?P<interval>[^-]+)-(?P<date>\d{4}-\d{2}-\d{2})$"
)


class YearlyKlineMerger:
    """Merge daily preprocessed CSV files into sorted yearly output files."""

    def __init__(self, input_folder: Path, output_folder: Path, overwrite: bool = False):
        self.input_folder = input_folder
        self.output_folder = output_folder
        self.overwrite = overwrite

    @staticmethod
    def _file_metadata(path: Path) -> dict[str, str]:
        match = FILENAME_PATTERN.match(path.stem)
        if not match:
            raise ValueError(f"Cannot read coin, interval, and date from filename: {path.name}")
        return match.groupdict()

    def _group_files(self) -> dict[tuple[str, str, str], list[Path]]:
        groups: dict[tuple[str, str, str], list[Path]] = defaultdict(list)
        for csv_path in self.input_folder.rglob("*.csv"):
            metadata = self._file_metadata(csv_path)
            key = (metadata["coin"], metadata["interval"], metadata["date"][:4])
            groups[key].append(csv_path)
        return groups

    @staticmethod
    def _excel_time(value: str) -> time:
        """Turn a preprocessed value such as 00:04:59.999Z into an Excel time."""
        return time.fromisoformat(value.rstrip("Z"))

    @staticmethod
    def _convert_value(column: str, value: str):
        if column == "date":
            return date.fromisoformat(value)
        if column in {"open_time", "close_time"}:
            return YearlyKlineMerger._excel_time(value)
        if column in {"coin", "interval"}:
            return value
        if column == "count":
            return int(value)
        if column == "ignore":
            return int(float(value))
        return float(value)

    def _write_workbook(self, csv_files: list[Path], output_path: Path) -> int:
        try:
            from openpyxl import Workbook
            from openpyxl.cell import WriteOnlyCell
            from openpyxl.styles import Font, PatternFill
        except ImportError as exc:
            raise RuntimeError(
                "Excel output needs openpyxl. Run: python -m pip install openpyxl"
            ) from exc

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("Klines")
        rows_written = 0
        header: list[str] | None = None

        for csv_path in sorted(csv_files):
            with csv_path.open("r", newline="", encoding="utf-8-sig") as source:
                reader = csv.DictReader(source)
                if reader.fieldnames is None:
                    raise ValueError(f"CSV has no header: {csv_path.name}")
                current_header = [name.strip() for name in reader.fieldnames]
                if header is None:
                    header = current_header
                    header_cells = []
                    for name in header:
                        cell = WriteOnlyCell(sheet, value=name)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor="1F4E78")
                        header_cells.append(cell)
                    sheet.append(header_cells)
                elif current_header != header:
                    raise ValueError(f"Columns do not match in {csv_path.name}")

                for row in reader:
                    excel_row = []
                    for column in header:
                        cell = WriteOnlyCell(sheet, value=self._convert_value(column, row[column]))
                        if column == "date":
                            cell.number_format = "yyyy-mm-dd"
                        elif column in {"open_time", "close_time"}:
                            cell.number_format = "hh:mm:ss.000"
                        excel_row.append(cell)
                    sheet.append(excel_row)
                    rows_written += 1

        if header is None:
            raise ValueError("No CSV data was found")

        workbook.save(output_path)
        return rows_written

    def _write_csv(self, csv_files: list[Path], output_path: Path) -> int:
        """Merge files into one standard CSV without changing their columns."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        expected_header: list[str] | None = None
        rows_written = 0

        with output_path.open("w", newline="", encoding="utf-8") as destination:
            writer: csv.DictWriter | None = None
            for csv_path in sorted(csv_files):
                with csv_path.open("r", newline="", encoding="utf-8-sig") as source:
                    reader = csv.DictReader(source)
                    if reader.fieldnames is None:
                        raise ValueError(f"CSV has no header: {csv_path.name}")
                    header = [name.strip() for name in reader.fieldnames]
                    if expected_header is None:
                        expected_header = header
                        writer = csv.DictWriter(destination, fieldnames=header)
                        writer.writeheader()
                    elif header != expected_header:
                        raise ValueError(f"Columns do not match in {csv_path.name}")

                    for row in reader:
                        writer.writerow(row)  # type: ignore[union-attr]
                        rows_written += 1
        return rows_written

    def _write_parquet(self, csv_files: list[Path], output_path: Path) -> int:
        """Merge files and write a compact, typed Parquet dataset."""
        try:
            import pandas as pd
        except ImportError as exc:
            raise RuntimeError(
                "Parquet output needs pandas and pyarrow. Run: python -m pip install pandas pyarrow"
            ) from exc

        frames = [pd.read_csv(path) for path in sorted(csv_files)]
        if not frames:
            raise ValueError("No CSV data was found")
        columns = list(frames[0].columns)
        if any(list(frame.columns) != columns for frame in frames[1:]):
            raise ValueError("Columns do not match across the input CSV files")

        data = pd.concat(frames, ignore_index=True)
        data["date"] = pd.to_datetime(data["date"], format="%Y-%m-%d")
        data.sort_values(["date", "open_time"], inplace=True, ignore_index=True)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            data.to_parquet(output_path, index=False, engine="pyarrow")
        except ImportError as exc:
            raise RuntimeError(
                "Parquet output needs pyarrow. Run: python -m pip install pyarrow"
            ) from exc
        return len(data)

    def merge_all(self, output_format: str) -> None:
        groups = self._group_files()
        if not groups:
            raise FileNotFoundError(f"No preprocessed CSV files found in: {self.input_folder.resolve()}")

        completed = skipped = 0
        for (coin, interval, year), csv_files in sorted(groups.items()):
            output_path = self.output_folder / f"{coin}-{interval}-{year}.{output_format}"
            if output_path.exists() and not self.overwrite:
                print(f"Skipped existing: {output_path.name}")
                skipped += 1
                continue

            writers = {
                "xlsx": self._write_workbook,
                "csv": self._write_csv,
                "parquet": self._write_parquet,
            }
            rows = writers[output_format](csv_files, output_path)
            print(f"Created {output_path.name}: {rows:,} rows from {len(csv_files)} daily files")
            completed += 1

        print(f"Finished. Created: {completed}; skipped: {skipped}")
        print(f"Yearly files: {self.output_folder.resolve()}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("data/pre_processed/splitted"),
        help="folder containing daily CSVs (default: data/pre_processed/splitted)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/pre_processed"),
        help="folder for merged yearly files (default: data/pre_processed)",
    )
    parser.add_argument(
        "--format",
        choices=("xlsx", "csv", "parquet"),
        default="parquet",
        help="output type; Parquet is recommended for an automated risk agent",
    )
    parser.add_argument("--overwrite", action="store_true", help="replace existing yearly output files")
    args = parser.parse_args()

    YearlyKlineMerger(args.input, args.output, args.overwrite).merge_all(args.format)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
